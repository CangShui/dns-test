#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import sys
import time
import ipaddress
import threading
import requests
import pandas as pd  # pip install pandas openpyxl
import dns.resolver  # pip install dnspython
from concurrent.futures import ThreadPoolExecutor, as_completed
from tqdm import tqdm  # pip install tqdm
from typing import Optional, Dict, Any, List
import openpyxl
from openpyxl.styles import PatternFill

# 可选：pip install maxminddb
try:
    import maxminddb
    MMDB_AVAILABLE = True
except ImportError:
    MMDB_AVAILABLE = False

# 默认 DNS 列表 URL
DEFAULT_URL = "https://public-dns.info/nameservers.txt"

# 默认测试域名
TEST_DOMAINS = [
    "google.com", "facebook.com", "amazon.com", "microsoft.com",
    "apple.com", "cloudflare.com", "alibaba.com", "baidu.com",
    "tencent.com", "netflix.com"
]

def 加载_ip_mmdb_db(mmdb_file: str = "ip.mmdb") -> Optional[maxminddb.Reader]:
    """加载MMDB数据库"""
    if not MMDB_AVAILABLE or not os.path.exists(mmdb_file):
        return None
    try:
        reader = maxminddb.open_database(mmdb_file)
        print(f"✅ 加载MMDB: {os.path.getsize(mmdb_file)/1024/1024:.1f}MB")
        return reader
    except:
        return None

# 全局MMDB（线程安全）
_ip_mmdb_reader = None
_ip_mmdb_lock = threading.Lock()

def 检查_google_ip(ip: str) -> bool:
    """仅按组织名判断Google归属"""
    global _ip_mmdb_reader
    
    if _ip_mmdb_reader is None:
        with _ip_mmdb_lock:
            if _ip_mmdb_reader is None:
                _ip_mmdb_reader = 加载_ip_mmdb_db()
    
    if _ip_mmdb_reader is None:
        return False
    
    try:
        response = _ip_mmdb_reader.get(ip)
        if not response:
            return False
        
        # 纯组织名模糊匹配
        org_fields = ['autonomous_system_organization', 'organization', 'isp']
        org_text = ' '.join(str(response.get(f, '') or '').lower() for f in org_fields)
        
        google_keywords = ['google', 'google llc', 'google cloud', 'google.com', 'alphabet', 'gcp']
        return any(keyword in org_text for keyword in google_keywords)
    except:
        return False

def 创建干净_resolver(dns_server: str, timeout: float = 3.0):
    """创建全新无缓存Resolver"""
    resolver = dns.resolver.Resolver(configure=False)  # 禁用系统配置
    resolver.nameservers = [dns_server]
    resolver.timeout = timeout
    resolver.lifetime = timeout
    resolver.cache = None  # 强制禁用缓存
    return resolver

def 终极污染检测(dns_server: str, benchmark_ips: List[str]) -> str:
    """🔥 5次独立无缓存验证"""
    #print(f"\n🔍 检测 {dns_server}")
    #print(f"基准IP: {benchmark_ips[:3]}")
    
    # 基准IP检查
    for i, ip in enumerate(benchmark_ips[:3]):
        if not 检查_google_ip(ip):
            #print(f"  ❌ 基准[{i+1}]污染: {ip}")
            return "已污染"
    
    # 5次全新无缓存解析
    #print("5次独立无缓存验证...")
    纯净次数 = 0
    
    for i in range(5):
        resolver = 创建干净_resolver(dns_server)
        try:
            start = time.perf_counter()
            answers = resolver.resolve("google.com", "A")
            ips = [str(rdata) for rdata in answers]
            latency = (time.perf_counter() - start) * 1000
            
            #print(f"  [{i+1}] {latency:.0f}ms: {ips[:2]}")
            
            if all(检查_google_ip(ip) for ip in ips):
                纯净次数 += 1
            else:
                #print(f"  ❌ [{i+1}]污染IP: {ips}")
                return "已污染"
        except Exception as e:
            #print(f"  ❌ [{i+1}]失败: {e}")
            return "已污染"
    
    result = "未污染" if 纯净次数 == 5 else "已污染"
    #print(f"✅ {纯净次数}/5 → {result}")
    return result

def 读取_dns列表(url: str):
    """读取DNS列表"""
    print(f"下载DNS列表: {url}")
    resp = requests.get(url, timeout=15)
    resp.raise_for_status()
    lines = resp.text.splitlines()
    dns_list = []
    
    for line in lines:
        line = line.strip()
        if not line: continue
        for part in line.split():
            try:
                ipaddress.ip_address(part)
                dns_list.append(part)
                break
            except ValueError:
                continue
    print(f"共{len(dns_list)}个DNS")
    return dns_list

def 按IP版本过滤(dns_list, mode: str):
    """IP版本过滤"""
    v4, v6 = [], []
    for ip_str in dns_list:
        try:
            ip = ipaddress.ip_address(ip_str)
            (v4 if isinstance(ip, ipaddress.IPv4Address) else v6).append(ip_str)
        except:
            continue
    return v4 if mode == "4" else v6 if mode == "6" else v4 + v6

_thread_local = threading.local()

def 获取_resolver(dns_server: str, timeout_sec: float):
    """线程本地Resolver"""
    if not hasattr(_thread_local, "resolvers"):
        _thread_local.resolvers = {}
    key = (dns_server, timeout_sec)
    if key not in _thread_local.resolvers:
        r = dns.resolver.Resolver()
        r.nameservers = [dns_server]
        r.timeout = timeout_sec
        r.lifetime = timeout_sec
        _thread_local.resolvers[key] = r
    return _thread_local.resolvers[key]

def 执行_dns查询(domain: str, dns_server: str, record_type: str, timeout_sec: float):
    """单次DNS查询"""
    start = time.perf_counter()
    try:
        resolver = 获取_resolver(dns_server, timeout_sec)
        answers = resolver.resolve(domain, record_type)
        ip_list = [str(rdata) for rdata in answers]
        return True, (time.perf_counter() - start) * 1000, ip_list
    except:
        return False, (time.perf_counter() - start) * 1000, []

def 测试单个dns(dns_server: str, domains, ip_mode: str, timeout_sec: float, 
                延迟下限_ms: float, 开启污染检查: bool):
    """测试单个DNS"""
    总次数, 成功次数, 延迟列表, domain_ips = 0, 0, [], {}
    
    try:
        是IPv4 = isinstance(ipaddress.ip_address(dns_server), ipaddress.IPv4Address)
    except:
        是IPv4 = True
    
    for domain in domains:
        rtype = "A" if ip_mode != "6" and (ip_mode == "4" or 是IPv4) else "AAAA"
        总次数 += 1
        ok, latency, ips = 执行_dns查询(domain, dns_server, rtype, timeout_sec)
        domain_ips[domain] = ips
        
        if latency is not None and latency < 延迟下限_ms:
            return None
        if ok:
            成功次数 += 1
            延迟列表.append(latency)
    
    if 总次数 == 0 or not 延迟列表:
        return None
    
    成功率 = 成功次数 / 总次数
    avg = sum(延迟列表) / len(延迟列表)
    min_d = min(延迟列表)
    max_d = max(延迟列表)
    
    污染状态 = "待检测" if 开启污染检查 and 成功率 > 0.4 else "未测试"
    
    return {
        "dns_server": dns_server, "成功率": 成功率,
        "平均延迟_ms": avg, "最小延迟_ms": min_d, "最大延迟_ms": max_d,
        "dns污染": 污染状态, "google_ips": domain_ips.get("google.com", [])
    }

def 设置_excel样式(output_file: str, 开启污染检查: bool):
    """Excel美化"""
    try:
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        for col in ws.columns:
            let = col[0].column_letter
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[let].width = min(max_len + 2, 55)
        
        if 开启污染检查:
            green = PatternFill(start_color="90EE90", fill_type="solid")
            for row in range(2, ws.max_row + 1):
                if ws.cell(row, 6).value == "未污染":
                    for c in range(1, ws.max_column + 1):
                        ws.cell(row, c).fill = green
        wb.save(output_file)
    except:
        pass

def main():
    print("DNS基准测试")
    
    # 1. DNS列表
    print("\n1) DNS列表:")
    print("1) 默认( 数据来自 public-dns.info )  2) 自定义")
    choice = input("选择(1/2): ").strip() or "1"
    url = input("URL: ").strip() if choice == "2" else DEFAULT_URL
    
    dns_list = 读取_dns列表(url)
    if not dns_list:
        sys.exit(1)
    
    # 2. IP模式
    print("\n2) 模式:")
    print("1)IPv4 2)IPv6 3)双栈")
    mode = input("选择(1/2/3): ").strip() or "1"
    ip_mode = {"1": "4", "2": "6", "3": "46"}[mode]
    dns_list = 按IP版本过滤(dns_list, ip_mode)
    print(f"筛选: {len(dns_list)}个DNS")
    
    # 3. 线程数
    print("\n3) 线程:")
    threads = int(input("线程数(1-4096,默认64): ").strip() or "64")
    threads = max(1, min(threads, 4096))
    
    # 4. 测试域名
    print("\n4) 域名:")
    n = int(input(f"数量(1-10,默认3): ").strip() or "3")
    n = max(1, min(n, len(TEST_DOMAINS)))
    test_domains = TEST_DOMAINS[:n]
    
    # 5. 延迟设置
    print("\n5) 延迟(ms):")
    min_delay = float(input("下限(默认10): ").strip() or "10")
    timeout_ms = float(input("超时(默认300): ").strip() or "300")
    per_query_timeout_sec = timeout_ms / 1000
    
    # 6. 污染检查
    #print("\n6) 污染检查:")
    #print("1)开启 2)关闭")
    #pollute = input("选择(1/2): ").strip() or "2"
    #开启污染检查 = pollute == "1"
    pollute = 1  # 默认开启污染检查
    
    # 7. 基准测试
    print("\n🔍 基准测试...")
    结果列表 = []
    start_all = time.perf_counter()
    
    with ThreadPoolExecutor(max_workers=threads) as executor:
        futures = [executor.submit(测试单个dns, dns, test_domains, ip_mode,
                                  per_query_timeout_sec, min_delay, pollute)
                  for dns in dns_list]
        
        with tqdm(total=len(dns_list), desc="基准测试", unit="DNS") as pbar:
            for future in as_completed(futures):
                try:
                    res = future.result()
                    if res:
                        结果列表.append(res)
                except:
                    pass
                pbar.update(1)
    
    print(f"\n✅ 基准完成: {len(结果列表)}/{len(dns_list)}有效 ({time.perf_counter()-start_all:.1f}s)")
    
    if not 结果列表:
        sys.exit(1)
    
    
    # 8. 污染检测
    if 开启污染检查:
        candidates = [r for r in 结果列表 if r["dns污染"] == "待检测"]
        if candidates:
            print(f"\n🔥 检测 {len(candidates)}个候选...")
            with ThreadPoolExecutor(max_workers=threads//4) as executor:  # 降低并发
                futures = {executor.submit(终极污染检测, r["dns_server"], r["google_ips"]): r
                          for r in candidates}
                with tqdm(total=len(candidates), desc="污染检测", unit="DNS") as pbar:
                    for future in as_completed(futures):
                        r = futures[future]
                        try:
                            r["dns污染"] = future.result()
                        except:
                            r["dns污染"] = "已污染"
                        pbar.update(1)
    
    # 9. 默认状态
    for r in 结果列表:
        if r["dns污染"] == "待检测":
            r["dns污染"] = "未测试"
    
    # 10. Excel导出
    df = pd.DataFrame(结果列表)
    sort_cols = ["成功率", "平均延迟_ms"]
    if 开启污染检查:
        df["dns污染"] = df["dns污染"].fillna("未测试")
        sort_cols.append("dns污染")
    
    df.sort_values(by=sort_cols, ascending=[False, True, False], na_position='last', inplace=True)
    
    cols = ["dns_server", "成功率", "平均延迟_ms", "最小延迟_ms", "最大延迟_ms"]
    if 开启污染检查:
        cols.append("dns污染")
    df = df[cols].rename(columns={
        "dns_server": "DNS服务器", "成功率": "成功率",
        "平均延迟_ms": "平均延迟(ms)", "最小延迟_ms": "最小延迟(ms)",
        "最大延迟_ms": "最大延迟(ms)", "dns污染": "DNS污染"
    })
    
    output_file = "测试结果.xlsx"
    df.to_excel(output_file, index=False)
    设置_excel样式(output_file, 开启污染检查)
    
    print(f"\n🎉 保存: {output_file}")
    best = df.iloc[0]
    #print(f"🏆 最佳: {best['DNS服务器']} ({best['成功率']:.1%}, {best['平均延迟(ms)']:.0f}ms)")
    

if __name__ == "__main__":
    main()
